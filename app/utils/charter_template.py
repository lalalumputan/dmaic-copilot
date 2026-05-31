# app/utils/charter_template.py
"""
Generate downloadable Excel template untuk Project Charter dan Benefit Estimate.
Extract data dari file yang diupload user.
"""
from __future__ import annotations
import io
from typing import Any, Dict, Optional

import pandas as pd


# ──────────────────────────────────────────────
# 1. GENERATE TEMPLATE
# ──────────────────────────────────────────────

def generate_charter_benefit_template() -> bytes:
    """
    Return bytes dari Excel file dengan 2 sheet:
    - Project Charter
    - Benefit Estimate
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_charter_sheet(writer)
        _write_benefit_sheet(writer)
    output.seek(0)
    return output.read()


def _write_charter_sheet(writer):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.book.create_sheet("Project Charter")

    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "PROJECT CHARTER"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")

    # Business Case
    ws.merge_cells("A2:F2")
    ws["A2"] = "BUSINESS CASE / PROJECT DESCRIPTION"
    ws["A2"].font = Font(bold=True)
    ws["A2"].fill = PatternFill("solid", fgColor="D6E4F0")
    ws.merge_cells("A3:F5")
    ws["A3"] = ""
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")

    # Problem & Goal
    ws["A6"] = "PROBLEM STATEMENT (SMART)"
    ws["A6"].font = Font(bold=True)
    ws["A6"].fill = PatternFill("solid", fgColor="D6E4F0")
    ws["D6"] = "GOAL STATEMENT (SMART)"
    ws["D6"].font = Font(bold=True)
    ws["D6"].fill = PatternFill("solid", fgColor="D6E4F0")
    ws.merge_cells("A7:C9")
    ws["A7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("D7:F9")
    ws["D7"].alignment = Alignment(wrap_text=True, vertical="top")

    # Resources header
    ws.merge_cells("A10:C10")
    ws["A10"] = "PROJECT RESOURCES"
    ws["A10"].font = Font(bold=True)
    ws["A10"].alignment = Alignment(horizontal="center")
    ws["A10"].fill = PatternFill("solid", fgColor="1E3A5F")
    ws["A10"].font = Font(bold=True, color="FFFFFF")

    ws.merge_cells("D10:F10")
    ws["D10"] = "TIMELINE"
    ws["D10"].font = Font(bold=True, color="FFFFFF")
    ws["D10"].alignment = Alignment(horizontal="center")
    ws["D10"].fill = PatternFill("solid", fgColor="1E3A5F")

    # Resource rows
    resource_rows = [
        ("Project Leader", ""),
        ("Project Champion", ""),
        ("Mentoring BB Name", ""),
        ("Process Owner", ""),
        ("Controlling", ""),
    ]
    for i, (label, val) in enumerate(resource_rows, start=11):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = val
        ws.merge_cells(f"B{i}:C{i}")

    # Team members
    ws["A16"] = "Team Members"
    ws["A16"].font = Font(bold=True)
    ws["B16"] = "Function"
    ws["B16"].font = Font(bold=True)
    for i in range(17, 22):
        ws[f"A{i}"] = ""
        ws[f"B{i}"] = ""
        ws.merge_cells(f"B{i}:C{i}")

    # Timeline rows
    timeline_rows = [
        ("Kick-off", "", ""),
        ("Define", "", ""),
        ("Measure", "", ""),
        ("Analyze", "", ""),
        ("Improve", "", ""),
        ("Implement", "", ""),
        ("Control", "", ""),
    ]
    ws["D11"] = "Milestone"
    ws["E11"] = "Goal (End Date)"
    ws["F11"] = "Actual"
    ws["D11"].font = Font(bold=True)
    ws["E11"].font = Font(bold=True)
    ws["F11"].font = Font(bold=True)
    for i, (label, goal, actual) in enumerate(timeline_rows, start=12):
        ws[f"D{i}"] = label
        ws[f"D{i}"].font = Font(bold=True)
        ws[f"E{i}"] = goal
        ws[f"F{i}"] = actual

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18


def _write_benefit_sheet(writer):
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = writer.book.create_sheet("Benefit Estimate")

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "BENEFIT ESTIMATE"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")

    # Benefit Potentials
    ws["A2"] = "BENEFIT POTENTIALS:"
    ws["A2"].font = Font(bold=True)
    ws.merge_cells("A3:C6")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["D2"] = "THE WAY HOW THE BENEFIT IS CALCULATED (AGREED WITH THE CONTROLLER):"
    ws["D2"].font = Font(bold=True)
    ws.merge_cells("D3:H6")
    ws["D3"].alignment = Alignment(wrap_text=True, vertical="top")

    # KPI Table header
    ws.merge_cells("A7:H7")
    ws["A7"] = "KEY PERFORMANCE INDICATORS (KPI):"
    ws["A7"].font = Font(bold=True)
    ws["A7"].fill = PatternFill("solid", fgColor="D6E4F0")

    kpi_headers = ["No.", "KPI", "Period of Measurement", "Unit",
                   "Baseline (BSL)", "Target", "Δ in %", "Agreed by Controlling"]
    for col, h in enumerate(kpi_headers, start=1):
        from openpyxl.utils import get_column_letter
        cell = ws[f"{get_column_letter(col)}8"]
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i in range(9, 15):
        ws[f"A{i}"] = i - 8

    # Assumptions
    ws.merge_cells("A15:H15")
    ws["A15"] = "Assumptions for the Way of Benefit Calculation:"
    ws["A15"].font = Font(bold=True)
    ws["A15"].fill = PatternFill("solid", fgColor="D6E4F0")
    for i in range(16, 21):
        ws[f"A{i}"] = i - 15

    # Soft benefits
    ws.merge_cells("A21:H21")
    ws["A21"] = "Soft Benefits:"
    ws["A21"].font = Font(bold=True)
    ws["A21"].fill = PatternFill("solid", fgColor="D6E4F0")
    for i in range(22, 25):
        ws[f"A{i}"] = ""

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 20


# ──────────────────────────────────────────────
# 2. EXTRACT FROM UPLOADED FILE
# ──────────────────────────────────────────────

def extract_charter_from_upload(uploaded_file) -> Dict[str, Any]:
    """
    Extract Project Charter data dari uploaded Excel file.
    Returns dict dengan field-field charter.
    """
    try:
        xl = pd.ExcelFile(uploaded_file)
        if "Project Charter" not in xl.sheet_names:
            return {"available": False, "error": "Sheet 'Project Charter' tidak ditemukan."}

        ws_df = pd.read_excel(uploaded_file, sheet_name="Project Charter", header=None)
        data = ws_df.fillna("").astype(str)

        def _cell(row, col):
            try:
                val = data.iloc[row, col]
                return val.strip() if isinstance(val, str) else ""
            except Exception:
                return ""

        # Business case = baris 2-4, kolom 0 (merged A3:F5)
        business_case = _cell(2, 0) or _cell(3, 0) or _cell(4, 0)

        # Problem = baris 6-8, kolom 0
        problem = _cell(6, 0) or _cell(7, 0) or _cell(8, 0)

        # Goal = baris 6-8, kolom 3
        goal = _cell(6, 3) or _cell(7, 3) or _cell(8, 3)

        # Resources baris 10-14, label di col 0, value di col 1
        def _resource(row_idx):
            return _cell(row_idx, 1)

        charter = {
            "available": True,
            "project_leader":  _resource(10),
            "champion":        _resource(11),
            "mentor_bb":       _resource(12),
            "process_owner":   _resource(13),
            "controller":      _resource(14),
            "business_case":   business_case,
            "problem_from_charter": problem,
            "goal_from_charter":    goal,
        }

        # Team members baris 16-20
        team = []
        for r in range(16, 21):
            name = _cell(r, 0)
            func = _cell(r, 1)
            if name:
                team.append({"name": name, "function": func})
        charter["team_members"] = team

        # Timeline baris 11-17, col D(3)=label, E(4)=goal, F(5)=actual
        timeline_labels = ["kickoff", "define_end", "measure_end", "analyze_end",
                           "improve_end", "implement_end", "control_end"]
        timeline = {}
        for i, label in enumerate(timeline_labels):
            goal_date = _cell(11 + i, 4)
            actual_date = _cell(11 + i, 5)
            if goal_date:
                timeline[label] = goal_date
        charter["timeline"] = timeline

        return charter

    except Exception as e:
        return {"available": False, "error": str(e)}


def extract_benefit_from_upload(uploaded_file) -> Dict[str, Any]:
    """
    Extract Benefit Estimate data dari uploaded Excel file.
    """
    try:
        xl = pd.ExcelFile(uploaded_file)
        if "Benefit Estimate" not in xl.sheet_names:
            return {"available": False, "error": "Sheet 'Benefit Estimate' tidak ditemukan."}

        ws_df = pd.read_excel(uploaded_file, sheet_name="Benefit Estimate", header=None)
        data = ws_df.fillna("").astype(str)

        def _cell(row, col):
            try:
                val = data.iloc[row, col]
                return val.strip() if isinstance(val, str) else ""
            except Exception:
                return ""

        benefit_potentials = _cell(2, 0) or _cell(3, 0) or _cell(4, 0) or _cell(5, 0)
        calc_method = _cell(2, 3) or _cell(3, 3) or _cell(4, 3) or _cell(5, 3)

        # KPI rows: baris 8-13 (row index), cols 0-7
        kpi_table = []
        for r in range(8, 14):
            no     = _cell(r, 0)
            kpi    = _cell(r, 1)
            period = _cell(r, 2)
            unit   = _cell(r, 3)
            bsl    = _cell(r, 4)
            target = _cell(r, 5)
            delta  = _cell(r, 6)
            agreed = _cell(r, 7)
            if kpi and kpi not in ("nan", ""):
                kpi_table.append({
                    "no": no, "kpi": kpi, "period": period,
                    "unit": unit, "baseline": bsl, "target": target,
                    "delta_pct": delta, "agreed_by_controlling": agreed,
                })

        # Assumptions: baris 15-19
        assumptions = []
        for r in range(15, 20):
            val = _cell(r, 0)
            if val and val not in ("nan", str(r - 14)):
                assumptions.append(val)

        # Soft benefits: baris 21-23
        soft_benefits = []
        for r in range(21, 24):
            val = _cell(r, 0)
            if val and val not in ("nan", ""):
                soft_benefits.append(val)

        return {
            "available": True,
            "benefit_potentials": benefit_potentials,
            "calculation_method": calc_method,
            "kpi_table": kpi_table,
            "assumptions": assumptions,
            "soft_benefits": soft_benefits,
        }

    except Exception as e:
        return {"available": False, "error": str(e)}